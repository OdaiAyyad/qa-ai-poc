import streamlit as st
from groq import Groq
import os
from dotenv import load_dotenv
import json
import html
import re
import csv
import sys
from datetime import datetime

# ---------------- PAGE CONFIG ----------------

st.set_page_config(
    page_title="QA AI Assistant",
    layout="wide"
)

# ---------------- GROQ CLIENT ----------------
load_dotenv()

client = Groq(
    api_key=os.getenv("GROQ_API_KEY")
)

MAX_ATTACHMENT_PREVIEW_ROWS = 10
MAX_ATTACHMENT_SUMMARY_CHARS = 12000
MAX_VALIDATION_ROWS = 500
VALIDATION_HISTORY_FILE = "chat_history/validation_runs.json"

CONSTRAINT_OPERATORS = [
    "equals",
    "not equals",
    "contains",
    "is not null",
    "is null",
    "greater than",
    "less than",
    "in list",
]


def ensure_workspace_packages():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    site_packages = os.path.join(project_root, "venv", "Lib", "site-packages")

    if os.path.isdir(site_packages) and site_packages not in sys.path:
        sys.path.insert(0, site_packages)


def get_ticket_topic(ticket_content):
    for line in ticket_content.splitlines():
        cleaned_line = line.strip()

        if cleaned_line:
            return cleaned_line

    return "General investigation"


def read_ticket_content(ticket_id):
    try:
        with open(f"tickets/{ticket_id}.txt", "r", encoding="utf-8") as file:
            return file.read()
    except:
        return "Ticket not found."


def get_short_ticket_title(ticket_content, max_words=5):
    topic = get_ticket_topic(ticket_content)
    cleaned_topic = re.sub(r"[^a-zA-Z0-9\u0600-\u06FF ]+", " ", topic)
    words = [word for word in cleaned_topic.split() if len(word) > 1]

    if not words:
        return "General Investigation"

    return " ".join(words[:max_words])


def get_ticket_history_label(ticket_id, ticket_content=None):
    if ticket_content is None:
        ticket_content = read_ticket_content(ticket_id)

    return f"{ticket_id} - {get_short_ticket_title(ticket_content)}"


def get_available_tickets():
    try:
        ticket_files = [
            file_name[:-4]
            for file_name in os.listdir("tickets")
            if file_name.lower().endswith(".txt")
        ]

        return sorted(ticket_files)
    except:
        return []


def get_ticket_attachments(ticket_id):
    try:
        ticket_key = ticket_id.lower()

        return sorted([
            file_name
            for file_name in os.listdir("attachments")
            if ticket_key in file_name.lower()
        ])
    except:
        return []


def normalize_ticket_id(ticket_id):
    cleaned_ticket_id = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(ticket_id).strip().upper())
    cleaned_ticket_id = cleaned_ticket_id.strip("-_")

    return cleaned_ticket_id


def sanitize_upload_filename(file_name):
    base_name = os.path.basename(file_name)
    base_name = re.sub(r"[\\/:*?\"<>|]+", "-", base_name)
    base_name = base_name.strip()

    return base_name or "attachment"


def build_ticket_file_content(title, business_requirements, validation_notes):
    sections = [title.strip()]

    if business_requirements.strip():
        sections.extend([
            "",
            "Business Requirements:",
            business_requirements.strip(),
        ])

    if validation_notes.strip():
        sections.extend([
            "",
            "Validation Notes:",
            validation_notes.strip(),
        ])

    return "\n".join(sections).strip() + "\n"


def save_new_ticket(ticket_id, title, business_requirements, validation_notes, uploaded_files):
    normalized_ticket_id = normalize_ticket_id(ticket_id)

    if not normalized_ticket_id:
        return False, "Ticket ID is required."

    if not title.strip() and not business_requirements.strip():
        return False, "Add a ticket title or business requirements."

    os.makedirs("tickets", exist_ok=True)
    os.makedirs("attachments", exist_ok=True)

    ticket_path = os.path.join("tickets", f"{normalized_ticket_id}.txt")
    ticket_content = build_ticket_file_content(
        title or normalized_ticket_id,
        business_requirements,
        validation_notes
    )

    with open(ticket_path, "w", encoding="utf-8") as file:
        file.write(ticket_content)

    saved_files = []

    for uploaded_file in uploaded_files or []:
        safe_name = sanitize_upload_filename(uploaded_file.name)

        if normalized_ticket_id.lower() not in safe_name.lower():
            safe_name = f"{normalized_ticket_id}-{safe_name}"

        attachment_path = os.path.join("attachments", safe_name)

        with open(attachment_path, "wb") as file:
            file.write(uploaded_file.getbuffer())

        saved_files.append(safe_name)

    st.cache_data.clear()

    return True, f"Saved {normalized_ticket_id} with {len(saved_files)} attachment(s)."


def get_interesting_columns(columns):
    keywords = [
        "id",
        "student",
        "mobile",
        "phone",
        "status",
        "discount",
        "scholarship",
        "order",
        "code",
    ]

    return [
        column
        for column in columns
        if any(keyword in str(column).lower() for keyword in keywords)
    ]


def summarize_dataframe(df, source_name):
    summary = [
        f"Source: {source_name}",
        f"Rows: {len(df)}",
        f"Columns: {', '.join(str(column) for column in df.columns)}",
    ]

    interesting_columns = get_interesting_columns(df.columns)

    if interesting_columns:
        summary.append(
            "Likely QA-relevant columns: "
            + ", ".join(str(column) for column in interesting_columns)
        )

        for column in interesting_columns[:6]:
            values = (
                df[column]
                .dropna()
                .astype(str)
                .str.strip()
            )
            values = values[values != ""].unique()[:8]

            if len(values):
                summary.append(
                    f"Sample values for {column}: {', '.join(values)}"
                )

    return "\n".join(summary)


def summarize_table(headers, rows, source_name, total_rows=None):
    headers = [str(header) if header is not None else "" for header in headers]
    rows = [
        ["" if value is None else str(value) for value in row]
        for row in rows
    ]
    row_count = total_rows if total_rows is not None else len(rows)
    summary = [
        f"Source: {source_name}",
        f"Rows: {row_count}",
        f"Columns: {', '.join(headers)}",
    ]
    interesting_columns = get_interesting_columns(headers)

    if interesting_columns:
        summary.append(
            "Likely QA-relevant columns: " + ", ".join(interesting_columns)
        )

        for column in interesting_columns[:6]:
            column_index = headers.index(column)
            values = []

            for row in rows:
                if column_index < len(row):
                    value = row[column_index].strip()

                    if value and value not in values:
                        values.append(value)

                if len(values) >= 8:
                    break

            if values:
                summary.append(
                    f"Sample values for {column}: {', '.join(values)}"
                )

    return "\n".join(summary)


def read_csv_fallback(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        headers = next(reader, [])
        rows = []

        for index, row in enumerate(reader):
            if index < 50:
                rows.append(row)

        total_rows = index + 1 if "index" in locals() else 0

    return summarize_table(headers, rows, "CSV", total_rows)


def read_excel_fallback(file_path):
    ensure_workspace_packages()

    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_blocks = []

    for sheet in workbook.worksheets:
        rows_iterator = sheet.iter_rows(values_only=True)
        headers = next(rows_iterator, [])
        rows = []

        for index, row in enumerate(rows_iterator):
            if index < 50:
                rows.append(list(row))

        total_rows = max((sheet.max_row or (len(rows) + 1)) - 1, 0)
        sheet_blocks.append(
            summarize_table(headers, rows, f"Sheet: {sheet.title}", total_rows)
        )

    workbook.close()
    return "\n\n".join(sheet_blocks)


def sanitize_sql_identifier(value):
    cleaned_value = re.sub(r"[^a-zA-Z0-9_]+", "_", str(value).strip().lower())
    cleaned_value = cleaned_value.strip("_")

    return cleaned_value or "parsed_attachment"


def get_sql_value(value):
    stripped_value = str(value).strip()

    if stripped_value == "":
        return "''"

    try:
        float(stripped_value)
        return stripped_value
    except ValueError:
        return "'" + stripped_value.replace("'", "''") + "'"


def normalize_cell_value(value):
    if value is None:
        return ""

    return str(value).strip()


def parse_number(value):
    cleaned_value = str(value).strip().replace(",", "")

    if cleaned_value.endswith("%"):
        cleaned_value = cleaned_value[:-1].strip()

    try:
        return float(cleaned_value)
    except ValueError:
        return None


def values_match(actual, expected):
    actual_number = parse_number(actual)
    expected_number = parse_number(expected)

    if actual_number is not None and expected_number is not None:
        return actual_number == expected_number

    return actual.lower() == expected.lower()


def get_normalized_row(row, width):
    padded_row = list(row) + [""] * max(0, width - len(row))

    return tuple(
        normalize_cell_value(value).strip().lower()
        for value in padded_row[:width]
    )


def update_duplicate_check(row, row_number, width, seen_rows, duplicate_samples):
    normalized_row = get_normalized_row(row, width)

    if not any(normalized_row):
        return 0

    if normalized_row in seen_rows:
        if len(duplicate_samples) < 5:
            duplicate_samples.append({
                "row": row_number,
                "duplicates row": seen_rows[normalized_row],
                "values": " | ".join(normalized_row[:8]),
            })

        return 1

    seen_rows[normalized_row] = row_number
    return 0


def read_csv_table(file_path, attachment):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        headers = [str(header) for header in next(reader, [])]
        rows = []
        total_rows = 0
        duplicate_count = 0
        duplicate_samples = []
        seen_rows = {}

        for row_number, row in enumerate(reader, start=2):
            total_rows += 1
            duplicate_count += update_duplicate_check(
                row,
                row_number,
                len(headers),
                seen_rows,
                duplicate_samples
            )

            if len(rows) < MAX_VALIDATION_ROWS:
                rows.append(row)

    return [{
        "id": f"{attachment}::CSV",
        "label": f"{attachment} / CSV",
        "file": attachment,
        "sheet": "CSV",
        "columns": headers,
        "rows": rows,
        "total_rows": total_rows,
        "duplicate_count": duplicate_count,
        "duplicate_samples": duplicate_samples,
    }]


def read_excel_tables(file_path, attachment):
    ensure_workspace_packages()

    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    tables = []

    for sheet in workbook.worksheets:
        rows_iterator = sheet.iter_rows(values_only=True)
        headers = [str(header) if header is not None else "" for header in next(rows_iterator, [])]
        rows = []
        total_rows = 0
        duplicate_count = 0
        duplicate_samples = []
        seen_rows = {}

        for row_number, row in enumerate(rows_iterator, start=2):
            total_rows += 1
            normalized_values = ["" if value is None else str(value) for value in row]
            duplicate_count += update_duplicate_check(
                normalized_values,
                row_number,
                len(headers),
                seen_rows,
                duplicate_samples
            )

            if len(rows) < MAX_VALIDATION_ROWS:
                rows.append(normalized_values)

        tables.append({
            "id": f"{attachment}::{sheet.title}",
            "label": f"{attachment} / {sheet.title}",
            "file": attachment,
            "sheet": sheet.title,
            "columns": headers,
            "rows": rows,
            "total_rows": total_rows,
            "duplicate_count": duplicate_count,
            "duplicate_samples": duplicate_samples,
        })

    workbook.close()
    return tables


@st.cache_data(show_spinner=False)
def parse_attachment_tables(attachments):
    tables = []
    unsupported_files = []
    parse_errors = []

    for attachment in attachments:
        file_path = os.path.join("attachments", attachment)
        extension = os.path.splitext(attachment)[1].lower()

        try:
            if extension == ".csv":
                tables.extend(read_csv_table(file_path, attachment))
            elif extension == ".xlsx":
                tables.extend(read_excel_tables(file_path, attachment))
            else:
                unsupported_files.append(attachment)
        except Exception as error:
            parse_errors.append(f"{attachment}: {error}")

    return tables, unsupported_files, parse_errors


def get_table_by_id(tables, table_id):
    for table in tables:
        if table["id"] == table_id:
            return table

    return None


def evaluate_constraint(row_value, operator, expected_value):
    actual = normalize_cell_value(row_value)
    expected = normalize_cell_value(expected_value)

    if operator == "equals":
        return values_match(actual, expected)
    if operator == "not equals":
        return not values_match(actual, expected)
    if operator == "contains":
        return expected.lower() in actual.lower()
    if operator == "is not null":
        return actual != ""
    if operator == "is null":
        return actual == ""
    if operator == "greater than":
        actual_number = parse_number(actual)
        expected_number = parse_number(expected)
        return actual_number is not None and expected_number is not None and actual_number > expected_number
    if operator == "less than":
        actual_number = parse_number(actual)
        expected_number = parse_number(expected)
        return actual_number is not None and expected_number is not None and actual_number < expected_number
    if operator == "in list":
        expected_values = [
            item.strip().lower()
            for item in expected.split(",")
            if item.strip()
        ]
        return actual.lower() in expected_values

    return False


def validate_constraint(table, column, operator, expected_value):
    if not table or column not in table["columns"]:
        return {
            "status": "Needs Review",
            "checked_rows": 0,
            "passed_rows": 0,
            "failed_rows": 0,
            "sample_failures": [],
            "sample_passes": [],
            "value_counts": [],
            "note": "Column was not found in the selected data.",
        }

    column_index = table["columns"].index(column)
    checked_rows = 0
    passed_rows = 0
    sample_failures = []
    sample_passes = []
    value_counts = {}

    for row_number, row in enumerate(table["rows"], start=2):
        checked_rows += 1
        row_value = row[column_index] if column_index < len(row) else ""
        normalized_value = normalize_cell_value(row_value)
        passed = evaluate_constraint(row_value, operator, expected_value)
        value_counts[normalized_value or "(blank)"] = (
            value_counts.get(normalized_value or "(blank)", 0) + 1
        )

        if passed:
            passed_rows += 1
            if len(sample_passes) < 5:
                sample_passes.append({
                    "row": row_number,
                    "value": normalized_value,
                })
        elif len(sample_failures) < 5:
            sample_failures.append({
                "row": row_number,
                "value": normalized_value,
            })

    failed_rows = checked_rows - passed_rows
    top_values = [
        {"value": value, "count": count}
        for value, count in sorted(
            value_counts.items(),
            key=lambda item: item[1],
            reverse=True
        )[:8]
    ]

    if checked_rows == 0:
        status = "Needs Review"
        note = "No data rows were available to validate."
    elif failed_rows == 0:
        status = "Passed"
        note = ""
    else:
        status = "Failed"
        note = ""

    return {
        "status": status,
        "checked_rows": checked_rows,
        "passed_rows": passed_rows,
        "failed_rows": failed_rows,
        "sample_failures": sample_failures,
        "sample_passes": sample_passes,
        "value_counts": top_values,
        "note": note,
    }


def generate_constraint_sql(table, column, operator, expected_value):
    table_name = sanitize_sql_identifier(f"{table['file']}_{table['sheet']}")
    column_name = sanitize_sql_identifier(column)
    expected_sql = get_sql_value(expected_value)

    if operator == "equals":
        where_clause = f"({column_name} <> {expected_sql} OR {column_name} IS NULL)"
    elif operator == "not equals":
        where_clause = f"{column_name} = {expected_sql}"
    elif operator == "contains":
        escaped_value = str(expected_value).replace("'", "''")
        where_clause = f"({column_name} NOT LIKE '%{escaped_value}%' OR {column_name} IS NULL)"
    elif operator == "is not null":
        where_clause = f"({column_name} IS NULL OR {column_name} = '')"
    elif operator == "is null":
        where_clause = f"{column_name} IS NOT NULL"
    elif operator == "greater than":
        where_clause = f"({column_name} <= {expected_sql} OR {column_name} IS NULL)"
    elif operator == "less than":
        where_clause = f"({column_name} >= {expected_sql} OR {column_name} IS NULL)"
    elif operator == "in list":
        values = [
            get_sql_value(item.strip())
            for item in str(expected_value).split(",")
            if item.strip()
        ]
        where_clause = f"{column_name} NOT IN ({', '.join(values)})"
    else:
        where_clause = "1 = 1"

    return "\n".join([
        "-- Suggested read-only validation query",
        "-- Replace table/column names with the real DB schema before execution.",
        "-- If the source file contains percentages, confirm how the DB stores them.",
        f"SELECT *",
        f"FROM {table_name}",
        f"WHERE {where_clause};",
    ])


def suggest_constraints_for_table(table):
    suggestions = []

    for column in table.get("columns", []):
        column_lower = str(column).lower()

        if "discount" in column_lower:
            suggestions.append((column, "equals", "50"))
        elif "status" in column_lower:
            suggestions.append((column, "equals", "Accepted"))
        elif "mobile" in column_lower or column_lower.endswith("id") or "_id" in column_lower:
            suggestions.append((column, "is not null", ""))

        if len(suggestions) >= 5:
            break

    return suggestions


def find_number_near_keywords(text, keywords):
    lowered_text = text.lower()

    for keyword in keywords:
        for match in re.finditer(re.escape(keyword), lowered_text):
            start = max(match.start() - 80, 0)
            end = min(match.end() + 80, len(text))
            window = text[start:end]
            number_match = re.search(r"(\d+(?:\.\d+)?)\s*%?", window)

            if number_match:
                return number_match.group(1)

    return None


def find_number_near_exact_phrases(text, phrases):
    lowered_text = text.lower()

    for phrase in phrases:
        phrase_lower = phrase.lower()

        for match in re.finditer(re.escape(phrase_lower), lowered_text):
            start = max(match.start() - 60, 0)
            end = min(match.end() + 60, len(text))
            window = text[start:end]
            number_match = re.search(r"(\d+(?:\.\d+)?)\s*%?", window)

            if number_match:
                return number_match.group(1)

    return None


def get_column_sample_values(table, column, limit=20):
    if not table or column not in table.get("columns", []):
        return []

    column_index = table["columns"].index(column)
    values = []

    for row in table.get("rows", []):
        value = normalize_cell_value(row[column_index] if column_index < len(row) else "")

        if value and value not in values:
            values.append(value)

        if len(values) >= limit:
            break

    return values


def is_boolean_like_values(values):
    if not values:
        return False

    boolean_values = {"true", "false", "yes", "no", "1", "0"}

    return all(str(value).strip().lower() in boolean_values for value in values)


def infer_column_constraint(column, ticket_content, table=None):
    column_lower = str(column).lower()
    ticket_lower = ticket_content.lower()

    if "gift" in column_lower:
        if any(phrase in ticket_lower for phrase in ["no gift", "without gift", "gift id didn't", "gift id did not", "gift id not"]):
            return ("is null", "", "Ticket says gift should not be linked")

    if "scholarship" in column_lower and column_lower.endswith("id"):
        return ("is not null", "", "Scholarship identifier should exist")

    if column_lower.endswith("id") or "_id" in column_lower or column_lower == "id":
        return ("is not null", "", "Identifier fields should not be missing")

    if "mobile" in column_lower or "phone" in column_lower:
        return ("is not null", "", "Contact identifier should not be missing")

    if "status" in column_lower:
        sample_values = get_column_sample_values(table, column)

        if is_boolean_like_values(sample_values):
            return ("equals", "True", "Boolean status flag should be true")

        status_values = ["accepted", "active", "cancelled", "canceled", "completed", "pending", "rejected"]

        for status in status_values:
            if status in ticket_lower:
                return ("equals", status.title(), "Status mentioned in ticket requirement")

        return ("is not null", "", "Status should not be missing")

    if "percentage" in column_lower or column_lower.endswith("_percent") or column_lower.endswith("_percentage"):
        expected_value = find_number_near_keywords(ticket_content, ["discount", "percentage"])

        if expected_value:
            return ("equals", expected_value, "Numeric value inferred from ticket requirement")

        return ("is not null", "", "Percentage column should not be missing")

    if "amount" in column_lower:
        expected_value = find_number_near_exact_phrases(
            ticket_content,
            [str(column), str(column).replace("_", " ")]
        )

        if expected_value:
            return ("equals", expected_value, "Exact amount requirement found in ticket")

        return ("is not null", "", "Amount column should not be missing")

    if "discount" in column_lower:
        expected_value = find_number_near_keywords(ticket_content, ["discount"])

        if expected_value:
            return ("equals", expected_value, "Discount value inferred from ticket requirement")

        return ("is not null", "", "Financial/value column should not be missing")

    return None


def infer_phase1_constraints(ticket_content, tables):
    constraints = []

    for table in tables:
        if table.get("duplicate_count", 0) > 0:
            continue

        if table.get("total_rows", 0) == 0 or not table.get("rows"):
            continue

        for column in table.get("columns", []):
            inferred_constraint = infer_column_constraint(column, ticket_content, table)

            if not inferred_constraint:
                continue

            operator, expected_value, reason = inferred_constraint
            constraints.append({
                "table_id": table["id"],
                "table_label": table["label"],
                "file": table["file"],
                "sheet": table["sheet"],
                "column": column,
                "operator": operator,
                "expected_value": expected_value,
                "reason": reason,
            })

            if len([
                item
                for item in constraints
                if item["table_id"] == table["id"]
            ]) >= 6:
                break

    return constraints


def build_phase1_data_report(ticket, ticket_content, environment, parsed_tables):
    validation_results = []
    sql_queries = []
    duplicate_results = []

    for table in parsed_tables:
        duplicate_count = table.get("duplicate_count", 0)
        duplicate_results.append({
            "table_label": table["label"],
            "status": "Failed" if duplicate_count else "Passed",
            "duplicate_count": duplicate_count,
            "sample_failures": table.get("duplicate_samples", []),
        })

    clean_tables = [
        table
        for table in parsed_tables
        if table.get("duplicate_count", 0) == 0
    ]
    inferred_constraints = infer_phase1_constraints(ticket_content, clean_tables)

    for constraint in inferred_constraints:
        table = get_table_by_id(clean_tables, constraint["table_id"])

        if table is None:
            continue

        result = validate_constraint(
            table,
            constraint["column"],
            constraint["operator"],
            constraint["expected_value"]
        )
        sql_query = generate_constraint_sql(
            table,
            constraint["column"],
            constraint["operator"],
            constraint["expected_value"]
        )
        validation_results.append({
            **constraint,
            **result,
            "sql": sql_query,
        })
        sql_queries.append(sql_query)

    failed_constraints = sum(
        1 for result in validation_results if result["status"] == "Failed"
    )
    review_constraints = sum(
        1 for result in validation_results if result["status"] == "Needs Review"
    )
    failed_duplicates = sum(
        1 for result in duplicate_results if result["status"] == "Failed"
    )
    passed_constraints = sum(
        1 for result in validation_results if result["status"] == "Passed"
    )

    if failed_constraints or failed_duplicates:
        run_status = "Failed"
    elif review_constraints:
        run_status = "Needs Review"
    else:
        run_status = "Passed"

    return {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "ticket": ticket,
        "topic": get_short_ticket_title(ticket_content),
        "environment": environment,
        "status": run_status,
        "run_type": "Phase 1 Data Report",
        "constraints": inferred_constraints,
        "results": validation_results,
        "duplicate_results": duplicate_results,
        "sql_queries": sql_queries,
        "passed_constraints": passed_constraints,
        "failed_constraints": failed_constraints + failed_duplicates,
        "review_constraints": review_constraints,
    }


def build_phase2_db_plan(ticket, ticket_topic, target_environment, source_run):
    source_queries = source_run.get("sql_queries", [])
    planned_queries = []

    for index, query in enumerate(source_queries, start=1):
        planned_queries.append({
            "query": query,
            "purpose": f"DB validation query {index}",
            "status": "Not Executed",
            "note": "DB/Metabase connection is not configured in this POC.",
        })

    return {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "ticket": ticket,
        "topic": ticket_topic,
        "environment": target_environment,
        "status": "Needs Review",
        "run_type": "Phase 2 DB/Metabase Plan",
        "execution_status": "Not Connected",
        "source_run_type": source_run.get("run_type", "Validation"),
        "constraints": source_run.get("constraints", []),
        "results": [],
        "duplicate_results": [],
        "planned_queries": planned_queries,
        "sql_queries": source_queries,
        "passed_constraints": 0,
        "failed_constraints": 0,
        "review_constraints": len(planned_queries),
    }


def infer_question_constraint(question, table):
    question_lower = question.lower()
    columns = table.get("columns", [])
    selected_column = None

    for column in columns:
        column_lower = str(column).lower()

        if column_lower in question_lower or column_lower.replace("_", " ") in question_lower:
            selected_column = column
            break

    if not selected_column:
        return None

    if any(phrase in question_lower for phrase in ["not null", "not missing", "available", "exists", "exist"]):
        return selected_column, "is not null", ""

    if any(phrase in question_lower for phrase in ["is null", "missing", "empty", "blank", "not exist"]):
        return selected_column, "is null", ""

    operator = "equals"

    if any(phrase in question_lower for phrase in ["not equal", "not equals", "different from"]):
        operator = "not equals"
    elif any(phrase in question_lower for phrase in ["greater than", "more than", "above"]):
        operator = "greater than"
    elif any(phrase in question_lower for phrase in ["less than", "below", "under"]):
        operator = "less than"
    elif "contain" in question_lower:
        operator = "contains"

    value_match = re.search(r"['\"]([^'\"]+)['\"]", question)

    if value_match:
        expected_value = value_match.group(1)
    else:
        value_match = re.search(r"(\d+(?:\.\d+)?\s*%?)", question)
        expected_value = value_match.group(1).strip() if value_match else ""

    if operator not in ["is null", "is not null"] and not expected_value:
        return selected_column, "is not null", ""

    return selected_column, operator, expected_value


def find_question_table(question, tables, selected_table):
    if infer_question_constraint(question, selected_table):
        return selected_table

    for table in tables:
        if table["id"] == selected_table["id"]:
            continue

        if infer_question_constraint(question, table):
            return table

    return None


def read_json_file(path, fallback):
    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except:
        return fallback


def write_json_file(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with open(path, "w", encoding="utf-8") as file:
        json.dump(data, file, indent=4, ensure_ascii=False)


def save_validation_run(run_entry):
    runs = read_json_file(VALIDATION_HISTORY_FILE, [])
    runs.append(run_entry)
    write_json_file(VALIDATION_HISTORY_FILE, runs)


def render_validation_history_sidebar():
    runs = read_json_file(VALIDATION_HISTORY_FILE, [])
    st.sidebar.markdown("## SQL History")

    if not runs:
        st.sidebar.caption("No validation runs yet.")
        return

    sorted_runs = sorted(
        runs,
        key=lambda run: parse_timestamp(run.get("timestamp", "")),
        reverse=True
    )
    grouped_runs = {}

    for run in sorted_runs:
        ticket_id = run.get("ticket", "Unknown Ticket")
        topic = run.get("topic") or get_topic_for_ticket(ticket_id)
        grouped_runs.setdefault(f"{ticket_id} - {topic}", []).append(run)

    for history_label, ticket_runs in list(grouped_runs.items())[:8]:
        with st.sidebar.expander(f"{history_label} ({len(ticket_runs)} runs)"):
            for index, run in enumerate(ticket_runs[:8], start=1):
                run_type = run.get("run_type", "Validation")
                label = (
                    f"{index}. {run_type}: {run.get('status', 'Status')} - "
                    f"{run.get('environment', 'Env')}"
                )

                if st.button(
                    label,
                    key=f"history_{history_label}_{run.get('timestamp', index)}",
                    use_container_width=True
                ):
                    run_type = run.get("run_type", "")

                    if run_type == "Phase 1 Data Report":
                        st.session_state.phase1_run = run
                        st.session_state.phase1_report_done = True
                    elif run_type == "Phase 2 DB/Metabase Plan":
                        st.session_state.phase2_run = run
                    elif run_type == "Natural Language SQL Check":
                        st.session_state.phase3_run = run
                    elif run_type == "Manual Constraint Check":
                        st.session_state.manual_run = run

                st.caption(format_timestamp(run.get("timestamp", "")))


def render_compact_run_result(run, title="Output"):
    st.markdown(f"#### {title}")

    status = run.get("status", "Status")
    if status == "Passed":
        status_class = "passed"
    elif status == "Needs Review":
        status_class = "review"
    else:
        status_class = "failed"
    st.markdown(
        f"""
        <div class="run-status {status_class}">
            {html.escape(run.get('run_type', 'Validation'))}: {html.escape(status)}
            | {run.get('passed_constraints', 0)} passed
            | {run.get('failed_constraints', 0)} failed
            | {run.get('review_constraints', 0)} needs review
        </div>
        """,
        unsafe_allow_html=True
    )
    st.caption(
        "POC result is based on loaded attachment data only. DB/Metabase execution is not connected yet."
    )

    summary_rows = []

    for duplicate_result in run.get("duplicate_results", []):
        summary_rows.append({
            "Check": "Duplicate rows",
            "Source": duplicate_result.get("table_label", "-"),
            "Status": duplicate_result.get("status", "-"),
            "Passed": "-" if duplicate_result.get("status") != "Passed" else "Yes",
            "Failed": duplicate_result.get("duplicate_count", 0),
        })

    for result in run.get("results", []):
        summary_rows.append({
            "Check": f"{result.get('column', '-')} {result.get('operator', '')} {result.get('expected_value', '')}".strip(),
            "Source": result.get("table_label", "-"),
            "Status": result.get("status", "-"),
            "Passed": result.get("passed_rows", 0),
            "Failed": result.get("failed_rows", 0),
            "Note": result.get("note", ""),
        })

    for planned_query in run.get("planned_queries", []):
        summary_rows.append({
            "Check": planned_query.get("purpose", "DB validation query"),
            "Source": run.get("environment", "-"),
            "Status": planned_query.get("status", "Not Executed"),
            "Passed": "-",
            "Failed": "-",
            "Note": planned_query.get("note", ""),
        })

    if summary_rows:
        st.dataframe(summary_rows, use_container_width=True, hide_index=True)

    with st.expander("Failed Evidence", expanded=False):
        has_evidence = False

        for duplicate_result in run.get("duplicate_results", []):
            if duplicate_result.get("status") != "Passed" and duplicate_result.get("sample_failures"):
                has_evidence = True
                st.markdown(f"**{duplicate_result.get('table_label', '-')}: duplicate rows**")
                st.dataframe(
                    duplicate_result["sample_failures"],
                    use_container_width=True,
                    hide_index=True
                )

        for result in run.get("results", []):
            if result.get("status") == "Passed":
                continue

            if result.get("reason"):
                st.caption(f"{result.get('column', '-')}: {result['reason']}")

            if result.get("value_counts"):
                has_evidence = True
                st.markdown(f"**{result.get('column', '-')}: actual values that caused the failure**")
                st.dataframe(result["value_counts"], use_container_width=True, hide_index=True)

            if result.get("sample_failures"):
                has_evidence = True
                st.markdown(f"**{result.get('column', '-')}: sample failed rows**")
                st.dataframe(result["sample_failures"], use_container_width=True, hide_index=True)

        if not has_evidence:
            st.caption("No failed evidence to show.")

    with st.expander("Generated SQL", expanded=False):
        if run.get("sql_queries"):
            for query in run["sql_queries"]:
                st.code(query, language="sql")
        else:
            st.caption("No SQL generated for this run.")


@st.cache_data(show_spinner=False)
def build_attachment_content_summary(attachments):
    ensure_workspace_packages()

    try:
        import pandas as pd
        pandas_error = None
    except Exception as error:
        pd = None
        pandas_error = error

    summary_blocks = []
    parsed_files = 0
    unsupported_files = []

    if pandas_error:
        summary_blocks.append(
            "Pandas could not be loaded in this environment, so the app used "
            "a lightweight fallback reader for Excel/CSV parsing."
        )

    for attachment in attachments:
        file_path = os.path.join("attachments", attachment)
        extension = os.path.splitext(attachment)[1].lower()

        try:
            if extension in [".xlsx", ".xls"]:
                if pd is not None:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_blocks = [f"File: {attachment}"]

                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        sheet_blocks.append(
                            summarize_dataframe(df, f"Sheet: {sheet_name}")
                        )

                    summary_blocks.append("\n\n".join(sheet_blocks))
                elif extension == ".xlsx":
                    summary_blocks.append(
                        "\n".join([
                            f"File: {attachment}",
                            read_excel_fallback(file_path),
                        ])
                    )
                else:
                    unsupported_files.append(attachment)
                    continue

                parsed_files += 1
            elif extension == ".csv":
                if pd is not None:
                    df = pd.read_csv(file_path)
                    summary_blocks.append(
                        "\n".join([
                            f"File: {attachment}",
                            summarize_dataframe(df, "CSV"),
                        ])
                    )
                else:
                    summary_blocks.append(
                        "\n".join([
                            f"File: {attachment}",
                            read_csv_fallback(file_path),
                        ])
                    )

                parsed_files += 1
            else:
                unsupported_files.append(attachment)
        except Exception as error:
            summary_blocks.append(
                f"File: {attachment}\nCould not parse attachment: {error}"
            )

    if unsupported_files:
        summary_blocks.append(
            "Unsupported attachments not parsed yet: "
            + ", ".join(unsupported_files)
        )

    summary = "\n\n---\n\n".join(summary_blocks).strip()

    if len(summary) > MAX_ATTACHMENT_SUMMARY_CHARS:
        summary = (
            summary[:MAX_ATTACHMENT_SUMMARY_CHARS]
            + "\n\n[Attachment summary truncated for prompt size.]"
        )

    return summary, parsed_files, len(unsupported_files)


def get_topic_for_ticket(ticket_id):
    try:
        with open(f"tickets/{ticket_id}.txt", "r", encoding="utf-8") as file:
            return get_ticket_topic(file.read())
    except:
        return "General investigation"


def build_history_groups(history):
    groups = {}
    topic_cache = {}

    for item in history:
        ticket_id = item.get("ticket", "Unknown ticket")
        topic = item.get("topic")

        if not topic:
            if ticket_id not in topic_cache:
                topic_cache[ticket_id] = get_topic_for_ticket(ticket_id)

            topic = topic_cache[ticket_id]

        group_key = f"{ticket_id}::{topic}"

        if group_key not in groups:
            groups[group_key] = {
                "ticket": ticket_id,
                "topic": topic,
                "items": []
            }

        groups[group_key]["items"].append(item)

    return groups


def parse_timestamp(timestamp):
    if not timestamp:
        return datetime.min

    try:
        return datetime.fromisoformat(timestamp)
    except ValueError:
        try:
            return datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            return datetime.min


def format_timestamp(timestamp):
    if not timestamp:
        return ""

    try:
        return parse_timestamp(timestamp).strftime("%d %b %Y, %I:%M %p")
    except ValueError:
        return timestamp


ANALYSIS_SECTIONS = [
    ("Affected Areas", ["Affected Areas"]),
    ("DB Checks", ["DB Checks", "Important DB Checks"]),
    ("Risky Dependencies", ["Risky Dependencies"]),
    ("Suggested Investigation", ["Suggested Investigation"]),
    ("Suggested Test Scenarios", ["Suggested Test Scenarios", "Test Scenarios"]),
    ("Suggested SQL Queries", ["Suggested SQL Queries", "SQL Queries"]),
    ("Regression Focus", ["Regression Focus"]),
]

SECTION_ICONS = {
    "Affected Areas": "🎯",
    "DB Checks": "🗄️",
    "Risky Dependencies": "⚠️",
    "Suggested Investigation": "🧭",
    "Suggested Test Scenarios": "🧪",
    "Suggested SQL Queries": "🧾",
    "Regression Focus": "✅",
}


def normalize_heading(heading):
    return heading.lower().replace("important ", "").strip()


def parse_analysis_sections(response):
    sections = {}
    current_section = None
    current_lines = []

    for line in response.splitlines():
        heading_match = re.match(r"^\s*#{1,3}\s+(.+?)\s*$", line)

        if heading_match:
            heading = heading_match.group(1).strip()
            normalized_heading = normalize_heading(heading)
            matched_title = None

            for title, aliases in ANALYSIS_SECTIONS:
                if normalized_heading in [normalize_heading(alias) for alias in aliases]:
                    matched_title = title
                    break

            if matched_title:
                if current_section:
                    sections[current_section] = "\n".join(current_lines).strip()

                current_section = matched_title
                current_lines = []
                continue

        if current_section:
            current_lines.append(line)

    if current_section:
        sections[current_section] = "\n".join(current_lines).strip()

    return sections


def render_inline_markdown(text):
    escaped_text = html.escape(text.strip())
    return re.sub(r"`([^`]+)`", r"<code>\1</code>", escaped_text)


def render_section_body(content):
    html_lines = []
    in_list = False
    in_code_block = False
    code_lines = []

    for line in content.splitlines():
        stripped_line = line.strip()

        if stripped_line.startswith("```"):
            if in_code_block:
                html_lines.append(
                    "<pre><code>"
                    + html.escape("\n".join(code_lines))
                    + "</code></pre>"
                )
                code_lines = []
                in_code_block = False
            else:
                if in_list:
                    html_lines.append("</ul>")
                    in_list = False

                in_code_block = True

            continue

        if in_code_block:
            code_lines.append(line)
            continue

        if not stripped_line:
            if in_list:
                html_lines.append("</ul>")
                in_list = False
            continue

        if stripped_line.startswith("- "):
            if not in_list:
                html_lines.append("<ul>")
                in_list = True

            html_lines.append(f"<li>{render_inline_markdown(stripped_line[2:])}</li>")
        else:
            if in_list:
                html_lines.append("</ul>")
                in_list = False

            html_lines.append(f"<p>{render_inline_markdown(stripped_line)}</p>")

    if in_list:
        html_lines.append("</ul>")

    if in_code_block:
        html_lines.append(
            "<pre><code>"
            + html.escape("\n".join(code_lines))
            + "</code></pre>"
        )

    return "\n".join(html_lines)


def render_analysis_card(title, content):
    st.markdown(
        f"""
        <section class="qa-analysis-card">
            <h4>{SECTION_ICONS.get(title, "•")} {html.escape(title)}</h4>
            <div class="qa-analysis-body">
                {render_section_body(content)}
            </div>
        </section>
        """,
        unsafe_allow_html=True
    )


def render_ai_analysis(response):
    sections = parse_analysis_sections(response)

    if not sections:
        st.markdown(response)
        return

    tab_config = [
        ("Impact", ["Affected Areas", "DB Checks", "Risky Dependencies"]),
        ("QA Scenarios", ["Suggested Test Scenarios", "Regression Focus"]),
        ("SQL", ["Suggested SQL Queries"]),
        ("Investigation", ["Suggested Investigation"]),
    ]
    visible_tabs = [
        (label, section_titles)
        for label, section_titles in tab_config
        if any(sections.get(section_title) for section_title in section_titles)
    ]

    if not visible_tabs:
        st.markdown(response)
        return

    tabs = st.tabs([label for label, _ in visible_tabs])

    for tab, (_, section_titles) in zip(tabs, visible_tabs):
        with tab:
            st.markdown('<div class="qa-analysis-grid">', unsafe_allow_html=True)

            for title in section_titles:
                content = sections.get(title)

                if content:
                    render_analysis_card(title, content)

            st.markdown("</div>", unsafe_allow_html=True)


def calculate_confidence_score(
    ticket_content,
    attachment_count,
    attachment_summary,
    has_historical_context
):
    score = 48
    ticket_length = len(ticket_content.strip())

    if ticket_content and ticket_content != "Ticket not found.":
        score += min(ticket_length // 80, 22)

    if attachment_count:
        score += 10

    if attachment_summary:
        score += 14

    if has_historical_context:
        score += 8

    return max(35, min(score, 95))


def get_data_sources_used(
    ticket_content,
    attachments,
    attachment_summary,
    has_historical_context
):
    sources = []

    if ticket_content and ticket_content != "Ticket not found.":
        sources.append("Jira Ticket")

    if any(
        os.path.splitext(attachment)[1].lower() in [".xlsx", ".xls", ".csv"]
        for attachment in attachments
    ):
        sources.append("Excel Attachment")

    if attachment_summary:
        sources.append("Parsed Document")

    if has_historical_context:
        sources.append("Historical Investigation")

    return sources


def render_grounding_panel(confidence_score, data_sources):
    source_items = "".join(
        f"<li>✓ {html.escape(source)}</li>"
        for source in data_sources
    )

    st.markdown(
        f"""
        <div class="grounding-grid">
            <section class="grounding-card confidence-card">
                <span>Confidence Level</span>
                <strong>{confidence_score}%</strong>
            </section>
            <section class="grounding-card sources-card">
                <span>Sources Used</span>
                <ul>{source_items}</ul>
            </section>
        </div>
        """,
        unsafe_allow_html=True
    )


# ---------------- SIDEBAR ----------------

st.sidebar.image("banner.png", use_container_width=True)
st.sidebar.title("QA SQL Validation Assistant")

# ---------------- MAIN UI ----------------

st.title("🔎 QA SQL Validation Assistant")

st.markdown(
    "Understand ticket data, define critical constraints, and generate SQL-style validation evidence."
)

st.markdown(
    """
    <style>
    html, body, [data-testid="stAppViewContainer"] {
        background: #0f172a;
    }

    [data-testid="stAppViewContainer"] {
        color: #f8fafc;
    }

    h1, h2, h3 {
        color: #e0f2fe;
    }

    [data-testid="stExpander"] {
        border: 1px solid rgba(56, 189, 248, 0.26);
        border-radius: 8px;
        background: rgba(15, 23, 42, 0.42);
    }

    .qa-analysis-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
        gap: 0.85rem;
        margin-top: 0.5rem;
    }

    .qa-analysis-card {
        border: 1px solid rgba(125, 211, 252, 0.34);
        border-radius: 8px;
        background: linear-gradient(180deg, rgba(30, 41, 59, 0.96), rgba(15, 23, 42, 0.96));
        box-shadow: 0 14px 28px rgba(2, 132, 199, 0.12);
        padding: 1rem;
        min-height: 10rem;
    }

    .qa-analysis-card h4 {
        color: #bae6fd;
        font-size: 1.08rem;
        margin: 0 0 0.65rem 0;
    }

    .qa-analysis-body {
        color: #e2e8f0;
        font-size: 1rem;
        line-height: 1.55;
    }

    .qa-analysis-body ul {
        margin: 0;
        padding-left: 1.1rem;
    }

    .qa-analysis-body li {
        margin-bottom: 0.45rem;
    }

    .qa-analysis-body code {
        background: rgba(14, 165, 233, 0.16);
        color: #e0f2fe;
        border-radius: 4px;
        padding: 0.08rem 0.28rem;
    }

    .qa-analysis-body pre {
        background: rgba(2, 6, 23, 0.68);
        border: 1px solid rgba(125, 211, 252, 0.22);
        border-radius: 8px;
        margin: 0.6rem 0 0 0;
        overflow-x: auto;
        padding: 0.75rem;
    }

    .qa-analysis-body pre code {
        background: transparent;
        color: #e0f2fe;
        padding: 0;
        white-space: pre;
    }

    .grounding-grid {
        display: grid;
        grid-template-columns: minmax(180px, 0.65fr) minmax(260px, 1.35fr);
        gap: 0.85rem;
        margin: 0.75rem 0 1rem 0;
    }

    .grounding-card {
        border: 1px solid rgba(125, 211, 252, 0.34);
        border-radius: 8px;
        background: rgba(14, 165, 233, 0.12);
        padding: 0.9rem 1rem;
    }

    .grounding-card span {
        color: #bae6fd;
        display: block;
        font-size: 0.92rem;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }

    .confidence-card strong {
        color: #f8fafc;
        font-size: 2rem;
        line-height: 1;
    }

    .sources-card ul {
        color: #e2e8f0;
        list-style: none;
        margin: 0;
        padding: 0;
    }

    .sources-card li {
        margin-bottom: 0.25rem;
    }

    div[data-testid="stButton"] > button {
        height: 3.2rem;
        width: 100%;
        white-space: normal;
        line-height: 1.2;
        padding: 0.55rem 0.75rem;
        border: 1px solid #7dd3fc;
        border-radius: 8px;
        background: linear-gradient(180deg, #e0f2fe 0%, #bae6fd 100%);
        color: #0f172a;
        font-weight: 700;
        box-shadow: 0 8px 18px rgba(14, 165, 233, 0.16);
    }

    div[data-testid="stButton"] > button:hover {
        border-color: #38bdf8;
        background: linear-gradient(180deg, #f0f9ff 0%, #7dd3fc 100%);
        color: #082f49;
    }

    div[data-testid="stButton"] > button[kind="primary"] {
        height: 3.35rem;
        border: 1px solid #22d3ee;
        background: linear-gradient(180deg, #0891b2 0%, #0e7490 100%);
        color: #ffffff;
        font-size: 1.02rem;
        box-shadow: 0 12px 24px rgba(8, 145, 178, 0.28);
    }

    div[data-testid="stButton"] > button[kind="primary"]:hover {
        border-color: #67e8f9;
        background: linear-gradient(180deg, #06b6d4 0%, #0891b2 100%);
        color: #ffffff;
    }

    section[data-testid="stSidebar"] {
        min-width: 22rem;
        border-right: 1px solid rgba(125, 211, 252, 0.2);
    }

    section[data-testid="stSidebar"] [data-testid="stExpander"] {
        background: rgba(30, 41, 59, 0.9);
    }

    section[data-testid="stSidebar"] [data-testid="stExpander"] details {
        padding: 0;
    }

    section[data-testid="stSidebar"] [data-testid="stExpander"] summary {
        font-size: 0.9rem;
        line-height: 1.25;
    }

    section[data-testid="stSidebar"] [data-testid="stImage"] {
        margin-bottom: 0.4rem;
    }

    section[data-testid="stSidebar"] .qa-analysis-grid {
        grid-template-columns: 1fr;
    }

    section[data-testid="stSidebar"] .qa-analysis-card {
        min-height: auto;
        padding: 0.75rem;
    }

    section[data-testid="stSidebar"] .qa-analysis-card h4 {
        font-size: 0.98rem;
    }

    section[data-testid="stSidebar"] .qa-analysis-body {
        font-size: 0.92rem;
    }

    section[data-testid="stSidebar"] .grounding-grid {
        grid-template-columns: 1fr;
    }

    section[data-testid="stSidebar"] .confidence-card strong {
        font-size: 1.45rem;
    }

    .attachment-summary {
        border: 1px solid rgba(125, 211, 252, 0.32);
        border-radius: 8px;
        background: rgba(14, 165, 233, 0.12);
        color: #e0f2fe;
        font-weight: 700;
        margin: 0.35rem 0 0.75rem 0;
        padding: 0.75rem 0.9rem;
    }

    .attachment-summary span {
        color: #7dd3fc;
    }

    .sidebar-context-card {
        border: 1px solid rgba(125, 211, 252, 0.26);
        border-radius: 8px;
        background: rgba(14, 165, 233, 0.1);
        color: #e0f2fe;
        margin: 0.5rem 0 0.75rem 0;
        padding: 0.75rem 0.85rem;
    }

    .sidebar-context-card strong {
        color: #7dd3fc;
    }

    .context-row {
        display: grid;
        grid-template-columns: 6.25rem 1fr;
        gap: 0.45rem;
        margin-bottom: 0.35rem;
    }

    .context-row:last-child {
        margin-bottom: 0;
    }

    .run-status {
        border-radius: 8px;
        font-weight: 700;
        margin: 0.35rem 0 0.75rem 0;
        padding: 0.65rem 0.8rem;
    }

    .run-status.passed {
        background: rgba(34, 197, 94, 0.16);
        border: 1px solid rgba(34, 197, 94, 0.38);
        color: #bbf7d0;
    }

    .run-status.failed {
        background: rgba(248, 113, 113, 0.14);
        border: 1px solid rgba(248, 113, 113, 0.34);
        color: #fecaca;
    }

    .run-status.review {
        background: rgba(250, 204, 21, 0.12);
        border: 1px solid rgba(250, 204, 21, 0.34);
        color: #fef3c7;
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.caption(
    "Flow: load ticket data, run the Phase 1 data report, add manual checks if needed, then ask data questions that generate SQL-style results."
)

with st.expander("Add New Ticket", expanded=False):
    st.caption(
        "Use this intake form to add a Jira ticket and its files without touching the project folders."
    )

    with st.form("new_ticket_form", clear_on_submit=False):
        intake_col1, intake_col2 = st.columns([0.35, 0.65])

        with intake_col1:
            new_ticket_id = st.text_input(
                "Ticket ID",
                placeholder="Example: STF-9001"
            )
            new_ticket_title = st.text_input(
                "Short title",
                placeholder="Example: Scholarship Migration"
            )

        with intake_col2:
            new_business_requirements = st.text_area(
                "Business requirements",
                placeholder=(
                    "Paste the ticket requirement here. Example: "
                    "discount_percentage must equal 50 for all listed students."
                ),
                height=120
            )

        new_validation_notes = st.text_area(
            "Validation notes",
            placeholder=(
                "Optional. Add known columns/rules, expected values, or anything QA should validate."
            ),
            height=80
        )
        new_attachments = st.file_uploader(
            "Ticket files",
            type=["csv", "xlsx", "txt"],
            accept_multiple_files=True
        )
        save_ticket_clicked = st.form_submit_button(
            "Save Ticket For Validation",
            type="primary",
            use_container_width=True
        )

    if save_ticket_clicked:
        saved, message = save_new_ticket(
            new_ticket_id,
            new_ticket_title,
            new_business_requirements,
            new_validation_notes,
            new_attachments
        )

        if saved:
            st.success(message)
            st.caption("Select the ticket from Ticket ID, then load its fields.")
        else:
            st.warning(message)

col1, col2 = st.columns(2)

with col1:
    environment = st.selectbox(
        "Environment",
        ["Preprod", "Production", "Staging"]
    )

with col2:
    available_tickets = get_available_tickets()

    if available_tickets:
        ticket = st.selectbox("Ticket ID", available_tickets)
    else:
        ticket = st.text_input("Ticket ID")
        st.warning("No ticket files found in the tickets folder.")

ticket_attachments = get_ticket_attachments(ticket)
ticket_content = read_ticket_content(ticket)
ticket_topic = get_short_ticket_title(ticket_content)
attachment_count = len(ticket_attachments)
attachment_summary = ""
parsed_attachment_count = 0
unsupported_attachment_count = 0

render_validation_history_sidebar()

st.sidebar.markdown("## Current Ticket")
st.sidebar.markdown(
    f"""
    <div class="sidebar-context-card">
        <div class="context-row"><strong>Ticket</strong><span>{html.escape(ticket)}</span></div>
        <div class="context-row"><strong>Title</strong><span>{html.escape(ticket_topic)}</span></div>
        <div class="context-row"><strong>Files</strong><span>{attachment_count}</span></div>
    </div>
    """,
    unsafe_allow_html=True
)

if (
    "attachment_summary_by_ticket" not in st.session_state
    or st.session_state.get("attachment_summary_ticket") != ticket
):
    st.session_state.attachment_summary_by_ticket = ""
    st.session_state.attachment_summary_ticket = ticket

if attachment_summary:
    st.session_state.attachment_summary_by_ticket = attachment_summary
else:
    attachment_summary = st.session_state.attachment_summary_by_ticket

if not ticket_attachments:
    st.sidebar.caption("No files for this ticket.")

if (
    "constraint_ticket" not in st.session_state
    or st.session_state.constraint_ticket != ticket
):
    st.session_state.constraint_ticket = ticket
    st.session_state.parsed_tables = []
    st.session_state.validation_constraints = []
    st.session_state.phase1_run = None
    st.session_state.phase2_run = None
    st.session_state.phase3_run = None
    st.session_state.manual_run = None
    st.session_state.phase1_report_done = False

st.markdown("### 1. Load Ticket Data")

if ticket_attachments:
    st.info(
        f"{ticket} has {attachment_count} attachment(s). Load the fields to see available sheets and columns."
    )
else:
    st.warning("No attachments are available for this ticket, so file-based validation cannot run yet.")

if ticket_attachments:
    if st.button("Load Attachment Fields", use_container_width=True):
        with st.spinner("Reading attachment tables..."):
            parsed_tables, unsupported_files, parse_errors = parse_attachment_tables(
                tuple(ticket_attachments)
            )

        st.session_state.parsed_tables = parsed_tables
        st.session_state.parsed_unsupported_files = unsupported_files
        st.session_state.parsed_errors = parse_errors
        st.session_state.validation_constraints = []
        st.session_state.phase1_run = None
        st.session_state.phase2_run = None
        st.session_state.phase3_run = None
        st.session_state.manual_run = None
        st.session_state.phase1_report_done = False

    if not st.session_state.parsed_tables:
        st.caption("Nothing is loaded yet. Click the button above to start.")

parsed_tables = st.session_state.get("parsed_tables", [])

if parsed_tables:
    st.markdown("### 2. Data Quality Check")

    overview_rows = []
    duplicate_tables = [
        table
        for table in parsed_tables
        if table.get("duplicate_count", 0) > 0
    ]
    clean_tables = [
        table
        for table in parsed_tables
        if table.get("duplicate_count", 0) == 0
    ]

    for table in parsed_tables:
        overview_rows.append({
            "File / sheet": table["label"],
            "Rows": table["total_rows"],
            "Columns": len(table["columns"]),
            "Duplicate rows": table.get("duplicate_count", 0),
            "Suggested fields": ", ".join(get_interesting_columns(table["columns"])[:5]) or "-"
        })

    with st.expander("Loaded file fields", expanded=True):
        st.dataframe(overview_rows, use_container_width=True, hide_index=True)

    st.markdown("### 3. Run Phase 1 Data Report")
    st.caption(
        "This report validates the ticket files against basic data quality and business rules inferred from the ticket text."
    )

    if st.button("Run Data Validation Report", type="primary", use_container_width=True):
        phase1_report = build_phase1_data_report(
            ticket,
            ticket_content,
            "File Validation",
            parsed_tables
        )
        save_validation_run(phase1_report)
        st.session_state.phase1_run = phase1_report
        st.session_state.phase1_report_done = True
        st.toast("Phase 1 data report saved.")

    if st.session_state.get("phase1_run"):
        render_compact_run_result(
            st.session_state.phase1_run,
            "Phase 1 Output"
        )

    if duplicate_tables:
        blocked_table_ids = {
            table["id"]
            for table in duplicate_tables
        }
        st.session_state.validation_constraints = [
            constraint
            for constraint in st.session_state.validation_constraints
            if constraint.get("table_id") not in blocked_table_ids
        ]
        st.warning(
            "Some files/sheets have duplicated rows and are blocked. "
            "You can still validate the clean files/sheets below."
        )

        for table in duplicate_tables:
            st.markdown(
                f"**{table['label']}** has "
                f"`{table.get('duplicate_count', 0)}` duplicated row(s)."
            )

            if table.get("duplicate_samples"):
                st.dataframe(
                    table["duplicate_samples"],
                    use_container_width=True,
                    hide_index=True
                )

    if clean_tables and not st.session_state.get("phase1_report_done", False):
        if duplicate_tables:
            st.success(f"{len(clean_tables)} clean file/sheet(s) are available.")
        else:
            st.success("No duplicated rows found.")

        st.caption("Run the Phase 1 Data Report to unlock manual checks and data questions.")

    if clean_tables and st.session_state.get("phase1_report_done", False):
        if duplicate_tables:
            st.success(f"{len(clean_tables)} clean file/sheet(s) are available for validation.")
        else:
            st.success("No duplicated rows found. You can continue to validation.")

        st.markdown("### Manual Checks")

        table_options = {
            table["label"]: table["id"]
            for table in clean_tables
        }

        builder_col1, builder_col2 = st.columns([1.2, 0.8])

        with builder_col1:
            selected_table_label = st.selectbox(
                "Parsed file / sheet",
                list(table_options.keys())
            )
            selected_table = get_table_by_id(
                clean_tables,
                table_options[selected_table_label]
            )
            selected_column = st.selectbox(
                "Column",
                selected_table["columns"]
            )

        with builder_col2:
            selected_operator = st.selectbox("Constraint", CONSTRAINT_OPERATORS)
            expected_value = st.text_input(
                "Expected value",
                disabled=selected_operator in ["is null", "is not null"],
                placeholder="Example: 50, Accepted, 100"
            )

        st.caption(
            f"Rows available in selected sheet: {selected_table['total_rows']}."
        )

        suggestions = suggest_constraints_for_table(selected_table)

        if suggestions:
            with st.expander("Quick constraint suggestions"):
                suggestion_cols = st.columns(min(3, len(suggestions)))

                for index, suggestion in enumerate(suggestions):
                    column, operator, value = suggestion

                    with suggestion_cols[index % len(suggestion_cols)]:
                        if st.button(
                            f"{column} {operator} {value}".strip(),
                            key=f"suggestion_{selected_table['id']}_{index}",
                            use_container_width=True
                        ):
                            st.session_state.validation_constraints.append({
                                "table_id": selected_table["id"],
                                "table_label": selected_table["label"],
                                "file": selected_table["file"],
                                "sheet": selected_table["sheet"],
                                "column": column,
                                "operator": operator,
                                "expected_value": value,
                            })

        add_disabled = selected_operator not in ["is null", "is not null"] and not expected_value.strip()

        if st.button(
            "Add Constraint",
            type="primary",
            use_container_width=True,
            disabled=add_disabled
        ):
            st.session_state.validation_constraints.append({
                "table_id": selected_table["id"],
                "table_label": selected_table["label"],
                "file": selected_table["file"],
                "sheet": selected_table["sheet"],
                "column": selected_column,
                "operator": selected_operator,
                "expected_value": "" if selected_operator in ["is null", "is not null"] else expected_value,
            })
    elif duplicate_tables:
        st.error(
            "All loaded files/sheets have duplicates, so there is nothing clean to validate yet."
        )

constraints = st.session_state.get("validation_constraints", [])
validation_tables = [
    table
    for table in parsed_tables
    if table.get("duplicate_count", 0) == 0
]

if constraints:
    st.markdown("### Run Manual Check")
    st.caption("These checks will be validated against the loaded file data and saved as SQL-style history.")

    for index, constraint in enumerate(constraints, start=1):
        st.markdown(
            f"{index}. `{constraint['table_label']}` | "
            f"`{constraint['column']}` **{constraint['operator']}** "
            f"`{constraint['expected_value']}`"
        )

    action_col1, action_col2 = st.columns([1, 1])

    with action_col1:
        clear_constraints = st.button("Clear Constraints", use_container_width=True)

    with action_col2:
        run_validation = st.button(
            "Run SQL Validation Check",
            type="primary",
            use_container_width=True
        )

    if clear_constraints:
        st.session_state.validation_constraints = []
        st.session_state.manual_run = None
        st.rerun()

    if run_validation:
        validation_results = []
        sql_queries = []

        for constraint in constraints:
            table = get_table_by_id(validation_tables, constraint["table_id"])

            if table is None:
                continue

            result = validate_constraint(
                table,
                constraint["column"],
                constraint["operator"],
                constraint["expected_value"]
            )
            sql_query = generate_constraint_sql(
                table,
                constraint["column"],
                constraint["operator"],
                constraint["expected_value"]
            )
            validation_results.append({
                **constraint,
                **result,
                "sql": sql_query,
            })
            sql_queries.append(sql_query)

        failed_constraints = sum(
            1 for result in validation_results if result["status"] == "Failed"
        )
        review_constraints = sum(
            1 for result in validation_results if result["status"] == "Needs Review"
        )
        passed_constraints = sum(
            1 for result in validation_results if result["status"] == "Passed"
        )

        if failed_constraints:
            run_status = "Failed"
        elif review_constraints:
            run_status = "Needs Review"
        else:
            run_status = "Passed"
        run_entry = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "ticket": ticket,
            "topic": ticket_topic,
            "environment": environment,
            "status": run_status,
            "run_type": "Manual Constraint Check",
            "constraints": constraints,
            "results": validation_results,
            "sql_queries": sql_queries,
            "passed_constraints": passed_constraints,
            "failed_constraints": failed_constraints,
            "review_constraints": review_constraints,
        }

        save_validation_run(run_entry)
        st.session_state.manual_run = run_entry
        st.toast("Manual validation check saved.")

if st.session_state.get("manual_run"):
    render_compact_run_result(
        st.session_state.manual_run,
        "Manual Check Output"
    )

latest_run = st.session_state.get("manual_run") or st.session_state.get("phase1_run")

if latest_run and st.session_state.get("phase1_report_done", False):
    st.markdown("### Phase 2: DB/Metabase Validation Plan")
    st.caption(
        "No DB/Metabase connection is configured yet. This step prepares the read-only queries that should run later in Preprod or Production."
    )
    phase2_environment = st.selectbox(
        "Target environment for planned DB validation",
        ["Preprod", "Production"],
        key="phase2_environment"
    )
    can_plan_db = bool(latest_run.get("sql_queries"))

    if st.button(
        "Create DB Validation Plan",
        use_container_width=True,
        disabled=not can_plan_db
    ):
        phase2_plan = build_phase2_db_plan(
            ticket,
            ticket_topic,
            phase2_environment,
            latest_run
        )
        save_validation_run(phase2_plan)
        st.session_state.phase2_run = phase2_plan
        st.toast("DB/Metabase validation plan saved.")

    if not can_plan_db:
        st.caption("No SQL is available from the latest run yet.")

    if st.session_state.get("phase2_run"):
        render_compact_run_result(
            st.session_state.phase2_run,
            "Phase 2 Output"
        )

if validation_tables and st.session_state.get("phase1_report_done", False):
    st.markdown("### Phase 3: Ask Data Question To Generate SQL")
    st.caption(
        "These options come from the clean sheets parsed from the selected ticket attachments. Ask about columns shown in Loaded file fields."
    )
    phase3_table_options = {
        table["label"]: table["id"]
        for table in validation_tables
    }
    phase3_table_label = st.selectbox(
        "File/sheet to query",
        list(phase3_table_options.keys()),
        key="phase3_table"
    )
    phase3_table = get_table_by_id(
        validation_tables,
        phase3_table_options[phase3_table_label]
    )

    if "data_question" not in st.session_state:
        st.session_state.data_question = ""

    data_question = st.text_area(
        "Data question",
        placeholder="Example: Is discount_percentage equal to 50 for all rows?",
        key="data_question"
    )

    if st.button("Convert To SQL And Run", type="primary", use_container_width=True):
        if not data_question.strip():
            st.warning("Please enter a data question.")
            st.stop()

        matched_phase3_table = find_question_table(
            data_question,
            validation_tables,
            phase3_table
        )

        if not matched_phase3_table:
            st.warning(
                "Could not find the mentioned column in the clean loaded sheets. "
                "Select a sheet that has this column or mention an exact column name."
            )
            st.stop()

        inferred_question_constraint = infer_question_constraint(
            data_question,
            matched_phase3_table
        )
        question_column, question_operator, question_expected_value = inferred_question_constraint
        question_result = validate_constraint(
            matched_phase3_table,
            question_column,
            question_operator,
            question_expected_value
        )
        question_sql = generate_constraint_sql(
            matched_phase3_table,
            question_column,
            question_operator,
            question_expected_value
        )
        question_constraint = {
            "table_id": matched_phase3_table["id"],
            "table_label": matched_phase3_table["label"],
            "file": matched_phase3_table["file"],
            "sheet": matched_phase3_table["sheet"],
            "column": question_column,
            "operator": question_operator,
            "expected_value": question_expected_value,
            "question": data_question,
        }
        question_run = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "ticket": ticket,
            "topic": ticket_topic,
            "environment": "File Validation",
            "status": question_result["status"],
            "run_type": "Natural Language SQL Check",
            "question": data_question,
            "constraints": [question_constraint],
            "results": [{
                **question_constraint,
                **question_result,
                "reason": f"Generated from natural-language data question using {matched_phase3_table['label']}",
                "sql": question_sql,
            }],
            "sql_queries": [question_sql],
            "passed_constraints": 1 if question_result["status"] == "Passed" else 0,
            "failed_constraints": 0 if question_result["status"] == "Passed" else 1,
        }
        save_validation_run(question_run)
        st.session_state.phase3_run = question_run
        st.toast("Data question converted to SQL and saved.")

    if st.session_state.get("phase3_run"):
        render_compact_run_result(
            st.session_state.phase3_run,
            "Phase 3 Output"
        )
elif parsed_tables and not st.session_state.get("phase1_report_done", False):
    st.caption("Ask Data Question unlocks after running the Phase 1 Data Report.")
